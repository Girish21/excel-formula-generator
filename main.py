import re
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.cell.read_only import ReadOnlyCell

# regex to capture patterns like 'Ace-SP&L'!C20, ('Ace-SP&L'!C$120*), C356, D$89
root_pattern_regex = r"('(?:(?:Ace-(?:SP&L|SBS|SCFS))'![A-Z]\$?(?:(?:[0-9])+))|(?:[A-Z]\$?(?:(?:[0-9])+)))(\+|\-|\*|\/)?"
# regex to extract (Ace-SP&L, 20) from 'Ace-SP&L'!C20, 439 from D$439
extract_regex = r"(?:(?:'(Ace-(?:SP&L|SBS|SCFS))'![A-Z]\$?(\d{1,}))|[A-Z]\$?(\d{1,}))"


"""
    method to recursively find the name of the row for the given excel coordinate
    @params
        extracted_expression - excel coordinate (eg: C40, F$567, 'Ace-SP&L'!C20)
        root_dict - ref to root dict to fetch row name
"""


def find_name(extracted_expression, root_dict):
    # check if extracted_expression is of type formula and can be splitted one more level, eg. =B$30
    if len(list(
            filter(None, list(re.findall(extract_regex, extracted_expression))))) == 0:
        # extracted_expression is the required string and it does not match the pattern, eg. Gross Sales
        return extracted_expression
    # extract the required details from the formula
    extracted_expression = list(
        filter(None, list(re.findall(extract_regex, extracted_expression)[0])))
    if len(extracted_expression) > 1:
        # if the expression is of type 'Ace-SP&L'!C$120, extracted_expression will be ['Ace-SP&L', 120]
        if int(extracted_expression[1]) in root_dict[extracted_expression[0]]:
            # if the pair is in dict, recursively call find_name to check if it is still of type formula
            return find_name(root_dict[extracted_expression[0]][int(extracted_expression[1])], root_dict)
        else:
            # fail-safe return
            return extracted_expression[1]
    else:
        # the expression is of type D$20, E3, etc. thus refering the SA_Ratios sheet
        if int(extracted_expression[0]) in root_dict['SA-Ratios']:
            # if the pair is in dict, recursively call find_name to check if it is still of type formula
            return find_name(root_dict['SA-Ratios'][int(extracted_expression[0])], root_dict)
        else:
            # fail-safe return
            return extracted_expression[0]


"""
    helper method to parse the formula and extract string matching patterns such as SUM( from =SUM(,
    365*AVERAGE( from =365*AVERAGE( and remove '='

    @params
        contents - list of parts of formula splitted according to regex
"""


def clean_formula(contents):
    # create a copy for non-destructive inplace edits of the array
    temp_contents = copy.deepcopy(contents)
    # find if the 0th index is of the pattern =SUM(, =365*AVERAGE(
    parsed_first_part = list(filter(None, re.findall(
        r"=((?:.*)?(?:SUM\(|AVERAGE\(|\())", temp_contents[0])))
    if len(parsed_first_part) > 0:
        # if found, replace the 0th index of the pattern with SUM(, or 365*AVERAGE(, etc. thus removing '='
        temp_contents[0] = parsed_first_part[0]
    else:
        # 0th index is just '=', pop it
        temp_contents.pop(0)
    return temp_contents


"""
    method to extract the target details from the given formula, parse it, 
    and find the corresponding name of the cell the part is refering to

    @params
        value - cell's value
        root_dict - ref to root_dict
"""


def format_formula(value, root_dict):
    # split the formula using root_pattern_regex, eg, "='Ace-SP&L'!C20*C45" to ['=', "'Ace-SP&L'!C20", '*', 'C45']
    temp_contents = list(filter(None, re.split(
        root_pattern_regex, value)))
    contents = clean_formula(temp_contents)
    # loop through the splitted formula array and find the name of each part of the formula
    for i in range(len(contents)):
        contents[i] = find_name(contents[i], root_dict)
    return contents


"""
    method to build the target parsed excel
    @params
        book - ref to the loaded excel book
        root_dict - ref to root dict
        dest_sheet_ref - ref to the target excel to write to
"""


def build_excel(book, root_dict, dest_sheet_ref):
    row = 1  # init start row index
    for cells in book.iter_rows(min_col=2, max_col=19, min_row=7):
        for cell in cells:
            # filtering empty and literal cells
            if type(cell) == ReadOnlyCell and cell.value != None and re.match(r'^-?\d+(?:\.\d+)?$', str(cell.value)) is None:
                # if it is a formula cell and a valid formula exist
                if cell.column != 2 and len(list(re.findall(root_pattern_regex, cell.value))) > 0:
                    contents = format_formula(cell.value, root_dict)
                    # join the splitted array with updated names and generate the formula and write into B column
                    dest_sheet_ref["B{}".format(
                        str(row))] = ' '.join(contents)
                    # row will be updated only if formula can be generated for the non highlighted row thus filtering literal and empty rows
                    row += 1
                    # break from looping through all the columns of a row if a formula is found
                    break
                else:
                    # check if the cell value is of the pattern of a formula or regular text
                    if len(list(re.findall(root_pattern_regex, cell.value))) > 0:
                        contents = format_formula(cell.value, root_dict)
                        dest_sheet_ref["A{}".format(
                            str(row))] = ' '.join(contents)
                    else:
                        dest_sheet_ref["A{}".format(str(row))] = cell.value


""" 
    method to generate dict for faster lookup
    @params
        book - ref to the loaded excel book
        root_dict - ref to root dict
        sheet_list - list of sheets to parse and generate lookup
        min_col
        max_col
        min_row
"""


def generate_dict(book, root_dict, sheet_list, min_col, max_col, min_row):
    for sheet in sheet_list:
        temp_dict = {}  # {[key: number]: string}
        for cell in book[sheet].iter_rows(min_col=min_col, max_col=max_col, min_row=min_row):
            # check to discard None values and empty cells
            if type(cell[0]) == ReadOnlyCell and cell[0].value != None:
                # remove unicode value \u00A0
                temp_dict[cell[0].row] = cell[0].value.strip()
        root_dict[sheet] = temp_dict


def main():
    # Open workbook to write to
    write_book = Workbook()
    active_write_book = write_book.active

    # Open the given file to parse
    read_book = load_workbook('NF-SA Template 160519.xlsx', read_only=True)

    # List of helper sheets needs to be parsed
    sheets = ['Ace-SP&L', 'Ace-SBS', 'Ace-SCFS']
    # Root dict to hold the parsed data
    root_dict = {}

    # generate the dict for lookup of row names
    generate_dict(read_book, root_dict, ['SA-Ratios'], 2, 2, 7)
    generate_dict(read_book, root_dict, sheets, 1, 1, 4)

    # build the target excel
    build_excel(read_book['SA-Ratios'], root_dict, active_write_book)

    write_book.save(filename="result.xlsx")


if __name__ == "__main__":
    main()
